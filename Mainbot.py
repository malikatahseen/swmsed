from openpyxl.descriptors.base import String
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN, MIN_ROW
import tagui as r
import pytesseract
import openpyxl
import pandas as pd
import time
from Mapping_loader import MappingFileLoader

mf1 = MappingFileLoader.fetch_Parse_param_mapping()
mf2 = MappingFileLoader.fetch_Product_Mapping()
count_mf2 = len(mf2)


# for i in range(0,count_mf2):
#     #print(mf2[i])
#     short_name = str(mf2[i]).split(':')[0]
#     long_name = str(mf2[i]).split(':')[1]
#     print(short_name,"short name")
#     print(long_name,"Long name")

name_retail = mf1.__getitem__(0)
date_of_sale = mf1.__getitem__(1)
existing_file_record_range = mf1.__getitem__(2)
new_file_record_range = mf1.__getitem__(3)
purchase_record_range = mf1.__getitem__(4)

name_retail_rows = name_retail.get('name.retail')
date_of_sale_rows = date_of_sale.get('date.of.sale')
existing_file_record_range_rows = existing_file_record_range.get('existing.file.record.range')
new_file_record_range_rows = new_file_record_range.get('new.file.record.range')
purchase_record_range_rows = purchase_record_range.get('purchase.record.range')

name_retail__st_row = name_retail_rows.split(',')[0]
name_retail__end_col = name_retail_rows.split(',')[1]
date_of_sale__st_row = date_of_sale_rows.split(',')[0]
date_of_sale__end_col = date_of_sale_rows.split(',')[1]
existing_file_record_range__st_row = existing_file_record_range_rows.split(',')[0]
existing_file_record_range__end_row = existing_file_record_range_rows.split(',')[1]
new_file_record_range__st_row = new_file_record_range_rows.split(',')[0]
new_file_record_range__end_row = new_file_record_range_rows.split(',')[1]
purchase_record_range__st_row = purchase_record_range_rows.split(',')[0]
purchase_record_range__end_row = purchase_record_range_rows.split(',')[1]
#print (existing_file_record_range__st_row ,'and===',existing_file_record_range__end_row)

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

ExcelFile_Path = r'D:\SMWSED\config\SMWSED_sep02_2021.xlsx'
Data_file = openpyxl.load_workbook(ExcelFile_Path)
Data_sheet = Data_file['SMWSED']

r.init(visual_automation = True)
r.url('https://exciseportal.py.gov.in/puduvaicalal1/')
r.wait(15)
# while r.popup('https://exciseportal.py.gov.in/puduvaicalal1/') == True:
r.type('//*[@id="txtUName"]', 'MUNIYANDI')
r.type('//*[@id="tpass"]', 'kannanSMW@272')
r.wait(0.2)
r.snap('//*[@id="img"]/form/table/tbody/tr[4]/td[2]/table/tbody/tr/td/table/tbody/tr/td/img', "cap.png")
TextFromImage = pytesseract.image_to_string(r'cap.png')
r.type('//*[@id="img"]/form/table/tbody/tr[5]/td[2]/input', TextFromImage)
r.wait(0.2)
r.click('LoginBtn.png')
r.wait(3)
# else:
#     if r.popup('https://www.google.co.in/') == False:
r.click('PermitBnt.png')
r.wait(0.2)
r.click('DealerForm.png')
r.wait(2)
Retailer_name = Data_sheet.cell(row=2, column=2).value
Date_of_sale = Data_sheet.cell(row=3, column=2).value
Date_of_sale_str = Date_of_sale.split(':')[0]
Date_of_sale_date = Date_of_sale.split(':')[1]

r.type("//input[@name='dateofsale']",Date_of_sale_date)
r.wait(0.2)
r.click('Sales_type.png')
r.wait(0.2)
r.click('FL2_reatail.png')
r.wait(0.2)

existing_file_list = []
new_file_record_list = []
purchase_record_list = []

# iterate through excel and display data
for i in range(int(existing_file_record_range__st_row), int(existing_file_record_range__end_row)+1):
    #print("\n")
    #print("Row ", i, " data :")
    existing_file_list_col = []
    for j in range(1, Data_sheet.max_column+1):
        cell_obj = Data_sheet.cell(row=i, column=j)
        existing_file_list_col.append(cell_obj.value)
        #print(cell_obj.value, "existing record column value")
    existing_file_list.append(existing_file_list_col)


for i in range(int(new_file_record_range__st_row), int(new_file_record_range__end_row)+1):
    #print("\n")
    #print("Row ", i, " data :")
    new_file_record_list_col = []
    for j in range(1, Data_sheet.max_column+1):
        cell_obj = Data_sheet.cell(row=i, column=j)
        new_file_record_list_col.append(cell_obj.value)
       # print(cell_obj.value, "new file column value")
    new_file_record_list.append(new_file_record_list_col)


for i in range(int(purchase_record_range__st_row), int(purchase_record_range__end_row)+1):
   # print("\n")
    #print("Row ", i, " data :")
    purchase_record_list_col = []
    for j in range(1, Data_sheet.max_column+1):
        cell_obj = Data_sheet.cell(row=i, column=j)
        purchase_record_list_col.append(cell_obj.value)
        #print(cell_obj.value, "purchase file column value")
    purchase_record_list.append(purchase_record_list_col)

count_pur = len(purchase_record_list)
count_ex = len(existing_file_list)
count_new = len(new_file_record_list)

for x in range(0, count_pur):
    val = purchase_record_list[x][1]
    v2 = purchase_record_list[x][2]
    v3 = purchase_record_list[x][3]
    for y in range(0, count_ex):
        val2 = existing_file_list[y][1]
        v4 = existing_file_list[y][2]
        v5 = existing_file_list[y][3]
        for i in range(0, count_mf2):
           if existing_file_list[y][1] == ((str(mf2[i]).split(':')[0])[2:-1]):
               existing_file_list[y][1] = (str(mf2[i]).split(':')[1][2:-1])
               print(existing_file_list[y][1])
               r.wait(10)
               r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[5]/td[6]/select',str(existing_file_list[y][1]))
               r.wait(5)
               r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[6]/td[2]/select',str(existing_file_list[y][2]))
               r.wait(5)
               r.type('//*[@id="exstock2"]/input','[clear]')
               r.wait(2)
               r.type('//*[@id="exstock2"]/input', str(existing_file_list[y][3]))
               r.wait(2)
               r.ask('Do you want to submit')
    for z in range(0, count_new):
        val3 = new_file_record_list[z][1]
        v6 = new_file_record_list[z][2]
        v7 = new_file_record_list[z][3]

        if ((str(purchase_record_list[x][1].strip()) == str(new_file_record_list[z][1].strip())) and (str(purchase_record_list[x][2]).strip() == str(new_file_record_list[z][2]).strip())):

            for i in range(0, count_mf2):
              if new_file_record_list[z][1] == ((str(mf2[i]).split(':')[0])[2:-1]):
                new_file_record_list[z][1] = (str(mf2[i]).split(':')[1][2:-1])
                print("new and perchase stock are equal")
                r.wait(10)
                r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[5]/td[6]/select',str(new_file_record_list[z][1]))
                r.wait(5)
                r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[6]/td[2]/select',str(new_file_record_list[z][2]))
                r.wait(5)
                r.type('//*[@id="newstock2"]/input','[clear]' )
                r.wait(5)
                r.type('//*[@id="newstock2"]/input', str(new_file_record_list[z][3]))
                r.wait(5)
                r.type('//*[@id="totpurchase2"]/input','[clear]')
                r.wait(5)
                r.type('//*[@id="totpurchase2"]/input', str(purchase_record_list[x][3]))
                r.wait(2)
                r.ask('Do you want to submit')

        else:
            for i in range(0, count_mf2):
              if new_file_record_list[z][1] == ((str(mf2[i]).split(':')[0])[2:-1]):
                new_file_record_list[z][1] = (str(mf2[i]).split(':')[1][2:-1])
                print("new and perchase stock are not equal")
                r.wait(10)
                r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[5]/td[6]/select',str(new_file_record_list[z][1]))
                r.wait(5)
                r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[6]/td[2]/select',str(new_file_record_list[z][2]))
                r.wait(5)
                r.type('//*[@id="newstock2"]/input','[clear]' )
                r.wait(5)
                r.type('//*[@id="newstock2"]/input', str(new_file_record_list[z][3]))
                r.wait(2)
                r.ask('Do you want to submit')
r.close()