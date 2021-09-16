import tagui as r
import pytesseract
import openpyxl
import time

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

ExcelFile_Path = r'C:\Users\MALIKA\Desktop\mapping file.xlsx'
Data_file = openpyxl.load_workbook(ExcelFile_Path)
Data_sheet = Data_file['Sheet1']

# r.init(visual_automation = True)
# r.url('https://exciseportal.py.gov.in/puduvaicalal1/')
# r.wait(15)
# r.type('//*[@id="txtUName"]', 'MUNIYANDI')
# r.type('//*[@id="tpass"]', 'kannanSMW@272')
# r.wait(0.2)
# r.snap('//*[@id="img"]/form/table/tbody/tr[4]/td[2]/table/tbody/tr/td/table/tbody/tr/td/img',"cap.png")
# TextFromImage = pytesseract.image_to_string(r'cap.png')
# r.type('//*[@id="img"]/form/table/tbody/tr[5]/td[2]/input', TextFromImage)
# r.wait(0.2)
# r.click('LoginBtn.png')
r.click('PermitBnt.png')
r.click('DealerForm.png')
for row_num,row_val in enumerate(Data_sheet.iter_rows(max_col=1,min_row=2,max_row=Data_sheet.max_row)):
    # r.click('calendar_icon.png')
    SDate = Data_sheet.cell(row=row_num+2,column=2).value
    print('Sale Date  == ',SDate)
    #r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[5]/td[2]/input',SDate)
    # r.select(SDate)
    # r.click('Sales_type.png')
    # r.click('FL2_reatail.png')
    Prod_ERP_Name = Data_sheet.cell(row=row_num+2,column=3).value
    Prod_GOV_Name = Data_sheet.cell(row=row_num+2,column=4).value
    # r.click('Brand_Name.png')
    # r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[5]/td[6]/select',Prod_ERP_Name)
    PACK_SIZE = Data_sheet.cell(row=row_num+2,column=5).value
    # r.click('Pack_size.png')
    # r.type('//*[@id="normalsale"]/td/fieldset/table/tbody/tr[6]/td[2]/select',PACK_SIZE)
    SALE_BTLS = Data_sheet.cell(row=row_num+2,column=6).value
    NEW_MRP = Data_sheet.cell(row=row_num+2,column=7).value
    SALE_AMT = Data_sheet.cell(row=row_num+2,column=8).value
    DUTY_AMT = Data_sheet.cell(row=row_num+2,column=9).value
    DUTY = Data_sheet.cell(row=row_num+2,column=10).value
    STATUS = Data_sheet.cell(row=row_num+2,column=11).value

    if Prod_ERP_Name == Prod_GOV_Name :
        print('Similar Names')
        print( Data_sheet.cell(row=row_num+2,column=3).value ,'equel to == ',Data_sheet.cell(row=row_num+2,column=4).value)

#r.close()